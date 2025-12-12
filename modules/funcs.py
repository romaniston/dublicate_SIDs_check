import time
import subprocess
import os.path

import openpyxl
from openpyxl.styles import PatternFill, Font


output_message_ws_no_ping = 'Workstation не пингуется. Local SID не получен.'
output_message_dont_get_local_sid = ('Не удалось получить Local SID. Возможно утилита '
                                     'запущена не от имени администратора.')


def table_var_assign():
    if compare_sids_table_exists():
        compare_sids_xlsx = openpyxl.open('compare_sids.xlsx')
        sheet = compare_sids_xlsx.active
        compare_sids_xlsx.save('compare_sids.xlsx')
        return compare_sids_xlsx, sheet
    else:
        compare_sids_xlsx = False
        return compare_sids_xlsx, None


def get_ws_names_and_discr():
    ws_list = []
    ws_list_discr = []
    with open("workstations.txt", "r") as file:
        not_first_str = False
        for string in file:
            if not_first_str:
                split_string = string.split()

                n = 0
                is_discr = False

                for l in string:
                    if l == "\t":
                        if is_discr == True:
                            break
                        is_discr = True
                    n += 1

                ws_list.append(split_string[0])
                ws_list_discr.append(string[n:])
            else:
                pass
                not_first_str = True
    return ws_list, ws_list_discr


def ping_or_not(ws_list, n):
    command = f'ping {ws_list[0 + n]}'
    try:
        result = subprocess.run(command, shell=True, capture_output=True, text=True, encoding='cp866', timeout=30)
        if 'число' in str(result).split():
            bool_var = True
        else:
            bool_var = False
    except subprocess.TimeoutExpired:
        print(f'Команда "{command}" превысила тайм-аут 30 секунд.')
        bool_var = False
    return bool_var


def get_one_domain_sid(ws_list, ws_num):
    command = 'cd c:\\ & cd pstools & psgetsid ' + str(ws_list[ws_num]) + '$'
    result = subprocess.run(command, shell=True, capture_output=True, text=True, encoding='cp866')
    get_sid = str(result).split(':')
    get_clear_domain_sid = str(get_sid[-1]).split('\\n')
    return get_clear_domain_sid[1]


def get_one_local_sid(ws_list, ws_num):
    command = 'cd c:\\ & cd pstools & psgetsid \\\\' + str(ws_list[ws_num])
    result = subprocess.run(command, shell=True, capture_output=True, text=True, encoding='cp866')
    time.sleep(5)
    get_sid = str(result).split(':')
    get_clear_local_sid = str(get_sid[-1]).split('\\n')
    return get_clear_local_sid[1]


# getting names, local and domain sids and writing them in "compare_sids.xlsx"
def get_names_and_sids(ws_list, sheet, compare_sids_xlsx, ws_list_discr):
    n = 0
    all_ws_data_list = [[None] * 5 for _ in range(len(ws_list))]

    for string in range(len(ws_list)):
        done_discr_fulled = ''
        done_discr = ''

        ws_name_output = f'Workstation #{1+n}'
        ping_or_not_bool = ping_or_not(ws_list, n)
        all_ws_data_list[n][0] = 1 + n
        all_ws_data_list[n][1] = ws_list[0 + n]
        # sheet[2 + (string)][0].value = 1 + n
        # sheet[2 + (string)][1].value = ws_list[0 + n]
        for done_discr in ws_list_discr[0 + n]:
            done_discr_fulled += ' ' + done_discr
        all_ws_data_list[n][2] = done_discr_fulled
        # sheet[2 + (string)][2].value = done_discr_fulled
        if not ping_or_not_bool:
            all_ws_data_list[n][3] = output_message_ws_no_ping
            # sheet[2 + (string)][3].value = output_message_ws_no_ping
            domain_sid = get_one_domain_sid(ws_list, 0 + n)
            yield ws_name_output, output_message_ws_no_ping, None, None, domain_sid
        else:
            local_sid = get_one_local_sid(ws_list, 0 + n)
            domain_sid = get_one_domain_sid(ws_list, 0 + n)
            if local_sid[0] != 'S':
                all_ws_data_list[n][3] = output_message_dont_get_local_sid
                # sheet[2 + (string)][3].value = output_message_dont_get_local_sid
                yield ws_name_output, output_message_dont_get_local_sid, None, None, domain_sid
            else:
                all_ws_data_list[n][3] = local_sid
                # sheet[2 + (string)][3].value = local_sid
                yield ws_name_output, None, None, local_sid, domain_sid
        all_ws_data_list[n][4] = domain_sid
        # sheet[2 + (string)][4].value = domain_sid
        n += 1

    # writing in "compare_sids.xlsx"
    num = -1
    for data in all_ws_data_list:
        num += 1
        for n2 in range(5):
            sheet[2 + (num)][n2].value = data[n2]
    compare_sids_xlsx.save('compare_sids.xlsx')


def excel_field_color(color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    return fill


def looking_for_dubles(sheet, compare_sids_xlsx, not_domain_sid):
        
    ws_count = 0
    n = 1
    while sheet[n][0].value != None:
        ws_count += 1
        n += 1

    if not_domain_sid:
        print('Идёт поиск дублей Local SID')
        domain_sid = 0
    else:
        print('Идёт поиск дублей Domain SID')
        domain_sid = 1

    is_header_field = True
    for string in range(ws_count):
        if is_header_field:
            is_header_field = False
            continue

        compared_ws_num_source = sheet[1 + string][0].value
        sid_val_source = sheet[1 + string][3 + domain_sid].value
        list_of_compared = []

        for compare_string in range(ws_count):
            compared_ws_num = sheet[1 + compare_string][0].value
            compared_ws_name = sheet[1 + compare_string][1].value
            sid_val = sheet[1 + compare_string][3 + domain_sid].value
            if compared_ws_num_source != compared_ws_num:
                if sid_val == sid_val_source:
                    if sid_val != output_message_ws_no_ping:
                        if sid_val != output_message_dont_get_local_sid:
                            list_of_compared.append(f'№{compared_ws_num} - {compared_ws_name}')
                        else:
                            pass

        local_sid_field = sheet[1 + string][3].value
        list_of_compared_done = ''
        if list_of_compared != []:
            for duble in list_of_compared:
                list_of_compared_done += str(duble) + ", "
            sheet[1 + string][5 + domain_sid].value = f'Найдены дубли: {list_of_compared_done}'
            sheet[1 + string][5 + domain_sid].fill = excel_field_color("FF0000")
        else:
            sheet[1 + string][5 + domain_sid].value = 'Совпадений нет'
            sheet[1 + string][5 + domain_sid].fill = excel_field_color("00FF00")

        if local_sid_field[0] != 'S':
            sheet[1 + string][5].value = 'Local SID не получен'
            sheet[1 + string][5].fill = excel_field_color("FF8000")
        else:
            pass

    try:
        compare_sids_xlsx.save('compare_sids.xlsx')
    except PermissionError:
        print('Нет доступа к таблице compare_sids.xlsx. Возможно, её нужно закрыть. Программа остановлена.')
        time.sleep(999999)

    if not_domain_sid:
        result = "Поиск дублей Local SID закончен"
    else:
        result = "Поиск дублей Domain SID закончен"

    return result


def create_compare_sids_table():
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    headers = [
        "№",
        "workstation", 
        "ws description",
        "local SID",
        "domain SID",
        "local SID dubles",
        "domain SID dubles"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    try:
        wb.save('compare_sids.xlsx')
        return True
    except PermissionError:
        print("Ошибка: Нет доступа к файлу. Закройте compare_sids.xlsx если он открыт.")
        return False
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    try:
        wb.save('compare_sids.xlsx')
        return True
    except PermissionError:
        print("Ошибка: Нет доступа к файлу. Закройте compare_sids.xlsx если он открыт.")
        return False


def compare_sids_table_exists():
    try:
        compare_sids_table = open('compare_sids.xlsx')
        print('Файл найден!')
        return True
    except Exception as e:
        print('Файл не найден. EXCEPTION: ' + str(e))
        return False
    

def get_missing_sids(ws_list, sheet, compare_sids_xlsx, ws_list_discr):
    n = 0
    all_ws_data_list = [[None] * 5 for _ in range(len(ws_list))]
    
    for i in range(len(ws_list)):
        all_ws_data_list[i][0] = sheet[2 + i][0].value  # №
        all_ws_data_list[i][1] = sheet[2 + i][1].value  # workstation
        all_ws_data_list[i][2] = sheet[2 + i][2].value  # ws description
        all_ws_data_list[i][3] = sheet[2 + i][3].value  # local SID
        all_ws_data_list[i][4] = sheet[2 + i][4].value  # domain SID

    for string in range(len(ws_list)):
        ws_name_output = f'Workstation #{1+n}'
        current_local_sid = all_ws_data_list[n][3]
        current_domain_sid = all_ws_data_list[n][4]
        
        # checking for update local sid
        need_update_local = (
            current_local_sid == output_message_dont_get_local_sid or
            current_local_sid == output_message_ws_no_ping or
            current_local_sid is None or
            str(current_local_sid).strip() == ''
        )
        
        # checking for update domain sid
        need_update_domain = (
            current_domain_sid is None or
            str(current_domain_sid).strip() == '' or
            not str(current_domain_sid).startswith('S-1-5-21')
        )
        
        if need_update_local or need_update_domain:
            ping_or_not_bool = ping_or_not(ws_list, n)
            
            if not ping_or_not_bool:
                # ws is no ping
                if need_update_local:
                    all_ws_data_list[n][3] = output_message_ws_no_ping
                if need_update_domain:
                    domain_sid = get_one_domain_sid(ws_list, 0 + n)
                    all_ws_data_list[n][4] = domain_sid
                yield ws_name_output, output_message_ws_no_ping, None, None, all_ws_data_list[n][4]
            else:
                # ws is ping, trying to get sid
                if need_update_local:
                    local_sid = get_one_local_sid(ws_list, 0 + n)
                    if local_sid[0] != 'S':
                        all_ws_data_list[n][3] = output_message_dont_get_local_sid
                    else:
                        all_ws_data_list[n][3] = local_sid
                
                if need_update_domain:
                    domain_sid = get_one_domain_sid(ws_list, 0 + n)
                    all_ws_data_list[n][4] = domain_sid
                
                yield ws_name_output, None, None, all_ws_data_list[n][3], all_ws_data_list[n][4]
        else:
            yield ws_name_output, None, "skip", current_local_sid, current_domain_sid
        
        n += 1

    # save
    num = -1
    for data in all_ws_data_list:
        num += 1
        for n2 in range(5):
            sheet[2 + (num)][n2].value = data[n2]
    compare_sids_xlsx.save('compare_sids.xlsx')


